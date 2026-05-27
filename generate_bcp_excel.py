from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Colour palette ──────────────────────────────────────────────
GREEN  = "C6EFCE"; YELLOW = "FFEB9C"; RED    = "FFC7CE"
HEADER = "1F4E79"; SUBHDR = "2E75B6"; LIGHT  = "D9E1F2"
WHITE  = "FFFFFF"; GRAY   = "F2F2F2"

def hfill(hex_): return PatternFill("solid", fgColor=hex_)
def hfont(hex_, bold=False, sz=11): return Font(color=hex_, bold=bold, size=sz)
def border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)
def wrap(): return Alignment(wrap_text=True, vertical="top")
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── Scenario data ────────────────────────────────────────────────
scenarios = [
    # (ID, Set, Company, Description, Level, LevelColor,
    #  BIA_func, RTO, RPO, FinImpact,
    #  RiskThreat, Likelihood, Severity,
    #  RecoveryStrategy,
    #  DRPlan,
    #  Roles,
    #  AlternateSite,
    #  CommPlan,
    #  IRSteps,
    #  Testing,
    #  Maintenance,
    #  AWSServices)
    ("A", "Set 1 – E-Commerce (ShopNow)", "ShopNow",
     "Office Wi-Fi Outage", "Low", GREEN,
     "Internal staff communication", "4 hours", "N/A", "Minimal",
     "ISP failure or router hardware fault", "High", "Low",
     "Staff switch to mobile hotspots or WFH; IT contacts ISP",
     "No IT failover; backup 4G routers on-site",
     "IT Helpdesk: diagnose & escalate; Team leads: notify WFH",
     "Staff work remotely via VPN",
     "IT sends Slack/email to all staff within 15 min",
     "Detect → check router → contact ISP → activate backup router",
     "Annual drill: simulate ISP outage, verify backup router",
     "Review ISP SLA annually; update router firmware quarterly",
     "N/A"),

    ("B", "Set 1 – E-Commerce (ShopNow)", "ShopNow",
     "Key Supplier Shipment Delay", "Medium", YELLOW,
     "Order fulfillment, customer delivery", "48 hours", "N/A", "~$200K/day",
     "Supplier bankruptcy, port strike, natural disaster", "Medium", "Medium",
     "Activate secondary supplier; prioritize high-value orders; notify customers",
     "Update inventory/OMS with new supplier lead times",
     "Procurement Mgr: contact secondary supplier in 4h; Ops Mgr: re-prioritize queue",
     "Secondary warehouse if primary fulfillment center affected",
     "Internal: Ops team in 2h; External: customers in 24h",
     "Detect delay → escalate to Ops Mgr → activate secondary supplier → update ETA",
     "Semi-annual tabletop: simulate supplier failure",
     "Review supplier contracts & secondary supplier list every 6 months",
     "N/A"),

    ("C", "Set 1 – E-Commerce (ShopNow)", "ShopNow",
     "Ransomware Attack on Production DB", "High", RED,
     "Entire platform (orders, payments, customer data)", "4 hours", "1 hour",
     "$5,600/min + regulatory fines + reputational damage",
     "Ransomware via phishing or unpatched vulnerability", "Medium", "Critical",
     "Isolate infected systems; restore from clean backup; failover to AWS DR; engage IR firm",
     "Hot standby multi-region (ap-southeast-1 → ap-east-1); hourly RDS/S3 snapshots; Route 53 auto-failover in 5 min; 30-day retention",
     "CISO: lead response; DR Lead: execute runbook; Legal: GDPR notification; CEO: comms approval",
     "Hot standby AWS region — full platform in 4h",
     "Internal: 30 min; Customers: status page 1h; Regulators: 72h (GDPR); Media: post legal review",
     "GuardDuty alert → isolate → failover → forensics → patch → failback → post-incident report",
     "Quarterly tabletop; bi-annual full DR drill; annual pen test",
     "Critical patches in 24h; weekly backup restore test; BCP review after each incident",
     "RDS, S3, Route 53, GuardDuty, AWS Backup"),

    ("D", "Set 2 – AWS Cloud (CloudBiz)", "CloudBiz",
     "AWS Lambda Throttling", "Low", GREEN,
     "Background jobs (reports, email notifications)", "1 hour", "N/A", "Minimal – jobs queued",
     "Traffic spike exceeds Lambda concurrency limit (default 1,000)", "Medium", "Low",
     "SQS absorbs excess requests; request concurrency limit increase via Service Quotas",
     "Lambda + SQS DLQ; CloudWatch alarm → SNS; reserved concurrency for critical functions",
     "On-call Engineer: acknowledge alarm, review metrics; Platform Lead: submit quota increase",
     "N/A – SQS buffers all requests",
     "CloudWatch → SNS → PagerDuty → on-call engineer; no customer comms unless SLA breached",
     "CloudWatch detects throttle → SNS alert → engineer reviews → increase concurrency → monitor backlog",
     "Quarterly load test to simulate throttling; verify DLQ alerts",
     "Review concurrency limits monthly; set CloudWatch alarm at 80% threshold",
     "Lambda, SQS, CloudWatch, SNS, Service Quotas"),

    ("E", "Set 2 – AWS Cloud (CloudBiz)", "CloudBiz",
     "Single Availability Zone (AZ) Outage", "Medium", YELLOW,
     "Web application, API, database", "15 minutes", "Near-zero", "~$84K/hour",
     "AWS AZ-level hardware or power failure", "Low", "Medium",
     "Multi-AZ architecture absorbs failure automatically; ALB re-routes; RDS promotes standby",
     "Multi-AZ warm standby; EC2 ASG spans 3 AZs; RDS Multi-AZ failover in 60–120s; ALB health checks deregister in 30s; EFS/S3 AZ-resilient",
     "On-call Engineer: monitor AWS Health + CloudWatch; Platform Lead: confirm recovery",
     "Remaining AZs in same region serve all traffic automatically",
     "Internal: AWS Health + PagerDuty; External: status page if >5 min degradation",
     "AWS Health alert → CloudWatch alarms → ALB re-routes → RDS failover → ASG replaces instances → confirm recovery",
     "Semi-annual chaos test via AWS Fault Injection Simulator; verify RDS failover time",
     "Ensure ASG spans min 2 AZs; review RDS Multi-AZ after infra changes; monthly Health Dashboard review",
     "EC2, RDS Multi-AZ, ALB, Auto Scaling, EFS, S3, AWS Health, CloudWatch"),

    ("F", "Set 2 – AWS Cloud (CloudBiz)", "CloudBiz",
     "Full AWS Region Outage", "High", RED,
     "Entire SaaS platform – all services unavailable", "1 hour", "15 minutes",
     "$5,600/min + SLA breach penalties + customer churn",
     "Large-scale AWS region failure (e.g., ap-southeast-1 power/network event)", "Very Low", "Critical",
     "Pre-provisioned warm standby in ap-northeast-1; DNS failover via Route 53; continuous data replication",
     "Warm standby cross-region; Aurora Global DB (lag <1s); S3 CRR; DynamoDB Global Tables; ASG pre-warmed at 20% → scales to 100%; Route 53 failover in 60s; CloudFront unaffected; gradual failback via weighted routing",
     "CTO: declare DR event; DR Lead: execute runbook; DB Lead: promote Aurora Global; DevOps Lead: scale ASG; Comms Lead: customer updates",
     "Warm standby in ap-northeast-1 (Tokyo) – full capacity in 30 min",
     "Internal: PagerDuty war room in 10 min; Customers: status page 15 min / email 30 min; Enterprise: account mgr calls 30 min; Regulators: if data residency SLA affected",
     "Health Dashboard → Route 53 failover → CTO declares DR → promote Aurora → scale ASG → confirm healthy → notify customers → monitor → failback via weighted routing → post-incident report",
     "Quarterly tabletop; bi-annual full region failover drill (FIS); annual DR audit",
     "Weekly Aurora replication lag review; monthly S3 CRR metrics; update runbook after infra changes; quarterly Route 53 health check review",
     "Route 53, Aurora Global, S3 CRR, DynamoDB Global Tables, EC2 ASG, CloudFront, AWS Fault Injection Simulator"),
]

# ── Sheet 1: Scenario Library ────────────────────────────────────
ws = wb.active
ws.title = "Scenario Library"

columns = [
    ("ID", 5), ("Example Set", 22), ("Company", 12), ("Scenario", 30),
    ("Critical Level", 14), ("Affected Function", 28), ("RTO", 12), ("RPO", 12),
    ("Financial Impact", 22), ("Threat", 30), ("Likelihood", 12), ("Severity", 12),
    ("Recovery Strategy", 40), ("DR Plan", 50), ("Roles & Responsibilities", 40),
    ("Alternate Site", 30), ("Communication Plan", 40), ("Incident Response Steps", 50),
    ("Testing", 40), ("Maintenance", 40), ("AWS Services", 35),
]

# Header row
for col_idx, (col_name, col_width) in enumerate(columns, 1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.fill = hfill(HEADER)
    cell.font = hfont(WHITE, bold=True, sz=11)
    cell.alignment = center()
    cell.border = border()
    ws.column_dimensions[get_column_letter(col_idx)].width = col_width

ws.row_dimensions[1].height = 30

# Data rows
level_color = {"Low": GREEN, "Medium": YELLOW, "High": RED}
for row_idx, s in enumerate(scenarios, 2):
    row_data = [
        s[0], s[1], s[2], s[3], s[4],
        s[6], s[7], s[8], s[9],
        s[10], s[11], s[12],
        s[13], s[14], s[15], s[16], s[17], s[18], s[19], s[20], s[21]
    ]
    bg = GRAY if row_idx % 2 == 0 else WHITE
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = wrap()
        cell.border = border()
        # Level column coloring
        if col_idx == 5:
            cell.fill = hfill(level_color.get(value, WHITE))
            cell.font = Font(bold=True, size=10)
            cell.alignment = center()
        else:
            cell.fill = hfill(bg)
            cell.font = Font(size=10)
    ws.row_dimensions[row_idx].height = 80

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

# ── Sheet 2: BCP Components Reference ───────────────────────────
ws2 = wb.create_sheet("BCP Components")
comp_headers = ["#", "Phase", "Component", "Description", "Key Data Point", "Source"]
comp_widths  = [5, 20, 28, 50, 50, 30]

for col_idx, (h, w) in enumerate(zip(comp_headers, comp_widths), 1):
    cell = ws2.cell(row=1, column=col_idx, value=h)
    cell.fill = hfill(HEADER)
    cell.font = hfont(WHITE, bold=True)
    cell.alignment = center()
    cell.border = border()
    ws2.column_dimensions[get_column_letter(col_idx)].width = w

components = [
    (1, "1 – Understand", "Business Impact Analysis", "Identifies critical functions; defines RTO & RPO", "60% of companies losing data shut down within 6 months", "Boston Computing Network"),
    (2, "1 – Understand", "Risk Assessment", "Identifies threats; evaluates likelihood & severity", "95% of breaches caused by human error", "IBM Security Report"),
    (3, "2 – Prepare", "Recovery Strategies", "Plans to restore critical functions within RTO/RPO", "2.5x more likely to close without a recovery strategy", "FEMA"),
    (4, "2 – Prepare", "DR Plan & Strategy", "IT-focused: failover, backup, restore, failback", "93% of companies without DR close within 1 year of major data disaster", "University of Texas"),
    (5, "2 – Prepare", "Roles & Responsibilities", "BCP/DR team structure; clear ownership per task", "—", "ISO 22301"),
    (6, "2 – Prepare", "Alternate Site / Work Arrangements", "Hot/Warm/Cold standby; remote work capability", "—", "NIST SP 800-34"),
    (7, "2 – Prepare", "Communication Plan", "Internal & external notifications; contact lists", "Poor comms increases recovery time by up to 50%", "Deloitte Crisis Management Survey"),
    (8, "3 – Respond", "Incident Response Plan", "Immediate actions; escalation; decision authority", "IR team saves avg $2.66M per breach", "IBM Cost of a Data Breach 2023"),
    (9, "4 – Sustain", "Testing & Exercises", "Tabletop, simulation, full DR drills", "Only 23% of orgs test BCP annually", "Disaster Recovery Journal"),
    (10, "4 – Sustain", "Plan Maintenance", "Scheduled reviews; version control; change mgmt", "Regular testers recover 2x faster", "IBM"),
]

phase_colors = {
    "1 – Understand": "D9E1F2",
    "2 – Prepare":    "E2EFDA",
    "3 – Respond":    "FFF2CC",
    "4 – Sustain":    "FCE4D6",
}

for row_idx, row in enumerate(components, 2):
    bg = phase_colors.get(row[1], WHITE)
    for col_idx, val in enumerate(row, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = hfill(bg)
        cell.font = Font(size=10)
        cell.alignment = wrap()
        cell.border = border()
    ws2.row_dimensions[row_idx].height = 45

ws2.row_dimensions[1].height = 28
ws2.freeze_panes = "A2"

# ── Sheet 3: DR Strategy Matrix ──────────────────────────────────
ws3 = wb.create_sheet("DR Strategy Matrix")
dr_headers = ["Strategy", "RTO", "RPO", "Relative Cost", "Best For", "AWS Implementation", "Scenario"]
dr_widths   = [22, 15, 15, 16, 35, 45, 12]

for col_idx, (h, w) in enumerate(zip(dr_headers, dr_widths), 1):
    cell = ws3.cell(row=1, column=col_idx, value=h)
    cell.fill = hfill(HEADER)
    cell.font = hfont(WHITE, bold=True)
    cell.alignment = center()
    cell.border = border()
    ws3.column_dimensions[get_column_letter(col_idx)].width = w

dr_strategies = [
    ("Backup & Restore",        "Hours",        "Hours",          "$  (Lowest)",  "Non-critical, cold data, archival",
     "S3 + AWS Backup + Glacier; restore EC2 from AMI",                          "D"),
    ("Pilot Light",             "30–60 min",    "Minutes",        "$$ (Low)",     "Core services; minimal always-on infra",
     "Minimal EC2/RDS running; scale up on event via ASG + Route 53",            "—"),
    ("Warm Standby",            "15–30 min",    "Seconds–Minutes","$$$ (Medium)", "Business-critical SaaS; balanced cost/RTO",
     "Scaled-down ASG in DR region; Aurora Global; S3 CRR; Route 53 failover",  "E, F"),
    ("Multi-Site Active/Active","Near-zero",    "Near-zero",      "$$$$ (Highest)","Mission-critical; financial; zero tolerance",
     "Multi-region ALB + Global Accelerator; DynamoDB Global Tables; Aurora Global active/active", "—"),
]

dr_colors = [LIGHT, WHITE, LIGHT, WHITE]
for row_idx, (row, bg) in enumerate(zip(dr_strategies, dr_colors), 2):
    for col_idx, val in enumerate(row, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = hfill(bg)
        cell.font = Font(size=10)
        cell.alignment = wrap()
        cell.border = border()
    ws3.row_dimensions[row_idx].height = 50

ws3.row_dimensions[1].height = 28

# ── Save ─────────────────────────────────────────────────────────
output = "/Users/viet/my-data/courses/bcp-scenario-library.xlsx"
wb.save(output)
print(f"Saved: {output}")
